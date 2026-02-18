"""
Celery tasks for export processing.
"""
import logging
from datetime import timedelta
from io import BytesIO
from time import monotonic

from celery import shared_task
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from apps.results.models import ResultSet, ElementResultsCache
from apps.results.services import ResultDataService
from apps.results.services.data_assembler import (
    assemble_table_rows,
    dataset_to_table_projection,
)
from apps.results.services.providers.common import sort_load_case_columns
from config.result_types import RESULT_TYPE_BASE_MAP
from core.logging import build_correlation_context
from .models import ExportJob
from .progress_events import (
    send_complete as send_export_complete,
    send_error as send_export_error,
    send_progress as send_export_progress,
)

logger = logging.getLogger(__name__)

EXPORT_PROGRESS_PERSIST_INTERVAL_SECONDS = 1.0
EXPORT_PROGRESS_PERSIST_PERCENT_STEP = 5


def _export_log_context(
    job: ExportJob,
    *,
    result_set_id: int | None = None,
    task_id: str | None = None,
) -> str:
    """Build a consistent log context for export lifecycle events."""
    resolved_result_set_id = (
        result_set_id if result_set_id is not None else (job.export_config or {}).get("result_set_id")
    )
    return build_correlation_context(
        project_id=job.project_id,
        project_slug=job.project.slug,
        job_type="export",
        job_id=job.id,
        result_set_id=resolved_result_set_id,
        task_id=task_id,
    )


def _get_export_table_rows(
    service: ResultDataService,
    result_set: ResultSet,
    result_type: str,
    direction: str,
    include_summary: bool,
):
    """
    Build export-ready table headers and rows from canonical result datasets.
    Returns None when no data exists for the result type/direction.
    """
    dataset = service.get_global_results(
        result_set_id=result_set.id,
        result_type=result_type,
        direction=direction,
    )
    if dataset is None:
        return None

    story_column = dataset.story_column
    try:
        projection = dataset_to_table_projection(
            dataset=dataset,
            include_summary=include_summary,
            fixed_columns=[story_column],
            required_columns=[story_column],
        )
        return projection["headers"], projection["rows"]
    except KeyError as exc:
        raise KeyError(f"{exc.args[0]} for {result_type}_{direction}") from exc


def _get_element_export_table_rows(
    service: ResultDataService,
    result_set: ResultSet,
    cache_type: str,
    base_type: str,
    include_summary: bool,
):
    """
    Build export rows for an element result type variant.
    For BeamRotations, uses the wide-format table data service.
    For other types, queries ElementResultsCache directly.
    Returns None when no data exists.
    """
    if base_type == "BeamRotations":
        table_data = service.get_beam_rotations_table_data(result_set.id)
        if table_data is None:
            return None

        fixed_cols = table_data["fixed_columns"]
        lc_cols = table_data["load_case_columns"]
        summary_cols = table_data["summary_columns"] if include_summary else []
        return assemble_table_rows(
            rows=table_data["rows"],
            fixed_columns=fixed_cols,
            load_case_columns=lc_cols,
            summary_columns=summary_cols,
        )

    cache_entries = (
        ElementResultsCache.objects.filter(
            project=service.project,
            result_set=result_set,
            result_type=cache_type,
        )
        .select_related("element", "story")
        .order_by("-story_sort_order", "element__name")
    )
    cache_entries = list(cache_entries)
    if not cache_entries:
        return None

    load_case_set = set()
    raw_rows = []

    for entry in cache_entries:
        row = {
            "Element": entry.element.name,
            "Story": entry.story.name,
        }
        for lc_name, value in entry.results_matrix.items():
            row[lc_name] = service._apply_multiplier(value, base_type)
            load_case_set.add(lc_name)

        if entry.avg_value is not None:
            row["Avg"] = service._apply_multiplier(entry.avg_value, base_type)
        if entry.max_value is not None:
            row["Max"] = service._apply_multiplier(entry.max_value, base_type)
        if entry.min_value is not None:
            row["Min"] = service._apply_multiplier(entry.min_value, base_type)

        raw_rows.append(row)

    lc_cols = sort_load_case_columns(list(load_case_set))
    summary_cols = (
        ["Avg", "Max", "Min"]
        if include_summary and any("Avg" in r for r in raw_rows)
        else []
    )
    return assemble_table_rows(
        rows=raw_rows,
        fixed_columns=["Element", "Story"],
        load_case_columns=lc_cols,
        summary_columns=summary_cols,
    )


def _get_joint_export_table_rows(
    service: ResultDataService,
    result_set: ResultSet,
    result_type: str,
    include_summary: bool,
):
    """
    Build export rows for a joint result type.
    Returns None when no data exists.
    """
    dataset = service.get_joint_results(
        result_set_id=result_set.id,
        result_type=result_type,
    )
    if dataset is None:
        return None

    projection = dataset_to_table_projection(
        dataset=dataset,
        include_summary=include_summary,
        fixed_columns=["Shell Object", "Unique Name"],
    )
    return projection["headers"], projection["rows"]


def _count_element_steps(element_types):
    """Count export steps for element types (accounting for variant expansion)."""
    count = 0
    for etype in element_types:
        base_map = RESULT_TYPE_BASE_MAP.get(etype, {})
        variants = base_map.get("variants", (etype,))
        count += len(variants)
    return count


def _should_persist_progress(
    *,
    current: int,
    total: int,
    last_persisted_current: int,
    last_persisted_at: float,
    now: float,
) -> bool:
    if current <= last_persisted_current:
        return False
    if current >= total:
        return True

    min_step_delta = max(int(total * EXPORT_PROGRESS_PERSIST_PERCENT_STEP / 100), 1)
    if (current - last_persisted_current) >= min_step_delta:
        return True

    return (now - last_persisted_at) >= EXPORT_PROGRESS_PERSIST_INTERVAL_SECONDS


def _write_excel_sheet(wb, sheet_name, headers, rows, header_fill, header_font, border):
    """Write a styled sheet to the workbook."""
    ws = wb.create_sheet(title=sheet_name[:31])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_values in enumerate(rows, 2):
        for col_idx, value in enumerate(row_values, 1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                continue
        ws.column_dimensions[column].width = min(max_length + 2, 20)

    return ws


@shared_task(bind=True)
def process_export_job(self, job_id: int):
    """Process export job and generate output file."""
    try:
        job = ExportJob.objects.get(id=job_id)
    except ExportJob.DoesNotExist:
        logger.error(f"Export job {job_id} not found")
        return

    try:
        job.status = "processing"
        config = dict(job.export_config or {})
        result_set_id = config.get("result_set_id")
        result_types = config.get("result_types", [])
        directions = config.get("directions", ["X", "Y"])
        element_types = config.get("element_types", [])
        joint_types = config.get("joint_types", [])
        include_summary = config.get("include_summary", True)
        global_steps = len(result_types) * len(directions)
        element_steps = _count_element_steps(element_types)
        joint_steps = len(joint_types)
        total_units = max(global_steps + element_steps + joint_steps, 1)

        config["progress_current"] = 0
        config["progress_total"] = total_units
        job.export_config = config
        job.save(update_fields=["status", "export_config"])
        logger.info(
            "Export task started (%s, format=%s, global_types=%s, element_types=%s, joint_types=%s)",
            _export_log_context(job, result_set_id=result_set_id, task_id=self.request.id),
            job.export_format,
            len(result_types),
            len(element_types),
            len(joint_types),
        )
        send_export_progress(
            job_id=job.id,
            message="Preparing export...",
            current=0,
            total=total_units,
        )

        progress_state = {
            "last_persisted_current": 0,
            "last_persisted_at": monotonic(),
        }

        def update_progress(current: int, total: int) -> None:
            config["progress_current"] = current
            config["progress_total"] = total
            now = monotonic()
            if _should_persist_progress(
                current=current,
                total=total,
                last_persisted_current=progress_state["last_persisted_current"],
                last_persisted_at=progress_state["last_persisted_at"],
                now=now,
            ):
                job.export_config = config
                job.save(update_fields=["export_config"])
                progress_state["last_persisted_current"] = current
                progress_state["last_persisted_at"] = now
            send_export_progress(
                job_id=job.id,
                message="Generating export...",
                current=current,
                total=total,
            )

        result_set = ResultSet.objects.get(id=result_set_id, project=job.project)

        if job.export_format == "excel":
            output = generate_excel_export(
                job.project,
                result_set,
                result_types,
                directions,
                include_summary,
                element_types=element_types,
                joint_types=joint_types,
                progress_callback=update_progress,
            )
            filename = f"{job.project.slug}_{result_set.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            job.output_file.save(filename, ContentFile(output.getvalue()))
            job.file_name = filename
            job.file_size = len(output.getvalue())

        elif job.export_format == "csv":
            output = generate_csv_export(
                job.project,
                result_set,
                result_types,
                directions,
                include_summary,
                element_types=element_types,
                joint_types=joint_types,
                progress_callback=update_progress,
            )
            filename = f"{job.project.slug}_{result_set.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            job.output_file.save(filename, ContentFile(output.encode("utf-8")))
            job.file_name = filename
            job.file_size = len(output.encode("utf-8"))

        job.status = "completed"
        job.completed_at = timezone.now()
        job.expires_at = timezone.now() + timedelta(hours=24)  # Files expire after 24h
        config["progress_current"] = total_units
        config["progress_total"] = total_units
        job.export_config = config
        job.save()
        send_export_complete(
            job_id=job.id,
            status="success",
            message="Export completed",
        )

        logger.info(
            "Export task completed (%s, file_name=%s, file_size=%s)",
            _export_log_context(job, result_set_id=result_set.id),
            job.file_name,
            job.file_size,
        )

    except (ExportJob.DoesNotExist, ResultSet.DoesNotExist, IOError) as e:
        logger.exception(
            "Export task failed (%s, error=%s)",
            _export_log_context(job),
            e,
        )
        job.status = "failed"
        job.error_message = str(e)
        job.save()
        send_export_error(job.id, "Export failed", str(e))
    except Exception:
        logger.exception("Unexpected export error (%s)", _export_log_context(job))
        job.status = "failed"
        job.error_message = "Unexpected export error"
        job.save()
        send_export_error(job.id, "Export failed", "Unexpected export error")
        raise


def generate_excel_export(
    project,
    result_set,
    result_types,
    directions,
    include_summary,
    element_types=None,
    joint_types=None,
    progress_callback=None,
):
    """Generate Excel export file."""
    wb = Workbook()
    global_steps = len(result_types) * len(directions)
    el_steps = _count_element_steps(element_types or [])
    jt_steps = len(joint_types or [])
    total_steps = max(global_steps + el_steps + jt_steps, 1)
    current_step = 0
    result_service = ResultDataService(project=project)

    # Style definitions
    header_fill = PatternFill(start_color="1c2128", end_color="1c2128", fill_type="solid")
    header_font = Font(bold=True, color="d1d5db")
    border = Border(bottom=Side(style="thin", color="2c313a"))

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Phase 1: Global results
    for result_type in result_types:
        for direction in directions:
            try:
                table_data = _get_export_table_rows(
                    service=result_service,
                    result_set=result_set,
                    result_type=result_type,
                    direction=direction,
                    include_summary=include_summary,
                )
                if table_data is None:
                    continue

                sheet_name = f"{result_type}_{direction}"
                headers, rows = table_data
                _write_excel_sheet(wb, sheet_name, headers, rows, header_fill, header_font, border)
            finally:
                current_step += 1
                if progress_callback:
                    progress_callback(current_step, total_steps)

    # Phase 2: Element results
    for etype in (element_types or []):
        base_map = RESULT_TYPE_BASE_MAP.get(etype, {})
        variants = base_map.get("variants", (etype,))

        for variant in variants:
            try:
                table_data = _get_element_export_table_rows(
                    service=result_service,
                    result_set=result_set,
                    cache_type=variant,
                    base_type=etype,
                    include_summary=include_summary,
                )
                if table_data is None:
                    continue

                headers, rows = table_data
                _write_excel_sheet(wb, variant, headers, rows, header_fill, header_font, border)
            finally:
                current_step += 1
                if progress_callback:
                    progress_callback(current_step, total_steps)

    # Phase 3: Joint results
    for jtype in (joint_types or []):
        try:
            table_data = _get_joint_export_table_rows(
                service=result_service,
                result_set=result_set,
                result_type=jtype,
                include_summary=include_summary,
            )
            if table_data is None:
                continue

            headers, rows = table_data
            _write_excel_sheet(wb, jtype, headers, rows, header_fill, header_font, border)
        finally:
            current_step += 1
            if progress_callback:
                progress_callback(current_step, total_steps)

    # If no sheets were created, add a placeholder
    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No data available for export")

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_csv_export(
    project,
    result_set,
    result_types,
    directions,
    include_summary,
    element_types=None,
    joint_types=None,
    progress_callback=None,
):
    """Generate CSV export (combines all data into single file)."""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)
    global_steps = len(result_types) * len(directions)
    el_steps = _count_element_steps(element_types or [])
    jt_steps = len(joint_types or [])
    total_steps = max(global_steps + el_steps + jt_steps, 1)
    current_step = 0
    result_service = ResultDataService(project=project)

    # Phase 1: Global results
    for result_type in result_types:
        for direction in directions:
            try:
                table_data = _get_export_table_rows(
                    service=result_service,
                    result_set=result_set,
                    result_type=result_type,
                    direction=direction,
                    include_summary=include_summary,
                )
                if table_data is None:
                    continue

                writer.writerow([])
                writer.writerow([f"--- {result_type} {direction} ---"])
                headers, rows = table_data
                writer.writerow(headers)
                for row in rows:
                    writer.writerow(["" if value is None else value for value in row])
            finally:
                current_step += 1
                if progress_callback:
                    progress_callback(current_step, total_steps)

    # Phase 2: Element results
    for etype in (element_types or []):
        base_map = RESULT_TYPE_BASE_MAP.get(etype, {})
        variants = base_map.get("variants", (etype,))

        for variant in variants:
            try:
                table_data = _get_element_export_table_rows(
                    service=result_service,
                    result_set=result_set,
                    cache_type=variant,
                    base_type=etype,
                    include_summary=include_summary,
                )
                if table_data is None:
                    continue

                writer.writerow([])
                writer.writerow([f"--- {variant} ---"])
                headers, rows = table_data
                writer.writerow(headers)
                for row in rows:
                    writer.writerow(["" if value is None else value for value in row])
            finally:
                current_step += 1
                if progress_callback:
                    progress_callback(current_step, total_steps)

    # Phase 3: Joint results
    for jtype in (joint_types or []):
        try:
            table_data = _get_joint_export_table_rows(
                service=result_service,
                result_set=result_set,
                result_type=jtype,
                include_summary=include_summary,
            )
            if table_data is None:
                continue

            writer.writerow([])
            writer.writerow([f"--- {jtype} ---"])
            headers, rows = table_data
            writer.writerow(headers)
            for row in rows:
                writer.writerow(["" if value is None else value for value in row])
        finally:
            current_step += 1
            if progress_callback:
                progress_callback(current_step, total_steps)

    return output.getvalue()
