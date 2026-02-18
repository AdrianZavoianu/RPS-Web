"""Tests for export tasks — table row builders, Excel/CSV generation, job processing."""

from io import BytesIO
from unittest.mock import patch, MagicMock

from django.contrib.auth import get_user_model
from django.test import TestCase

from apps.catalog.models import CatalogProject
from apps.projects.models import Element, LoadCase, Project, Story
from apps.results.models import (
    ElementResultsCache,
    GlobalResultsCache,
    JointResultsCache,
    ResultSet,
)
from apps.exporter.models import ExportJob
from apps.exporter.tasks import (
    _get_export_table_rows,
    _get_element_export_table_rows,
    _get_joint_export_table_rows,
    generate_csv_export,
    generate_excel_export,
    process_export_job,
)
from apps.results.services.result_service import ResultDataService

User = get_user_model()


class _ExportTaskTestMixin:
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(
            username="exp-task-user",
            email="exp-task@example.com",
            password="ExpTaskPass!123",
        )
        cls.catalog = CatalogProject.objects.create(
            name="Export Task Project",
            slug="export-task-project",
            owner=cls.user,
            analysis_type="NLTHA",
        )
        cls.project = Project.objects.create(catalog_project=cls.catalog)
        cls.story = Story.objects.create(project=cls.project, name="L1", sort_order=1)
        cls.lc1 = LoadCase.objects.create(project=cls.project, name="TH01", case_type="Time History")
        cls.lc2 = LoadCase.objects.create(project=cls.project, name="TH02", case_type="Time History")

        cls.result_set = ResultSet.objects.create(
            project=cls.project, name="RS-EXP", analysis_type="NLTHA"
        )
        cls.empty_result_set = ResultSet.objects.create(
            project=cls.project, name="RS-EXP-EMPTY", analysis_type="NLTHA"
        )

        GlobalResultsCache.objects.create(
            project=cls.project,
            result_set=cls.result_set,
            result_type="Drifts_X",
            story=cls.story,
            results_matrix={"TH01": 0.003, "TH02": 0.004},
            avg_value=0.0035,
            max_value=0.004,
            min_value=0.003,
            load_case_count=2,
            story_sort_order=1,
        )


# ------------------------------------------------------------------
# _get_export_table_rows (global)
# ------------------------------------------------------------------


class GetExportTableRowsTests(_ExportTaskTestMixin, TestCase):
    def test_headers_and_rows_valid(self):
        service = ResultDataService(self.project)
        result = _get_export_table_rows(
            service, self.result_set, "Drifts", "X", include_summary=True
        )
        self.assertIsNotNone(result)
        headers, rows = result
        self.assertIn("Story", headers)
        self.assertTrue(len(rows) > 0)

    def test_none_when_empty(self):
        service = ResultDataService(self.project)
        result = _get_export_table_rows(
            service, self.empty_result_set, "Drifts", "X", include_summary=True
        )
        self.assertIsNone(result)

    def test_summary_toggle(self):
        service = ResultDataService(self.project)
        with_summary = _get_export_table_rows(
            service, self.result_set, "Drifts", "X", include_summary=True
        )
        without_summary = _get_export_table_rows(
            service, self.result_set, "Drifts", "X", include_summary=False
        )
        h_with, _ = with_summary
        h_without, _ = without_summary
        self.assertIn("Avg", h_with)
        self.assertNotIn("Avg", h_without)


# ------------------------------------------------------------------
# _get_element_export_table_rows
# ------------------------------------------------------------------


class GetElementExportTableRowsTests(_ExportTaskTestMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.wall = Element.objects.create(
            project=cls.project, element_type="Wall", name="W1", unique_name="W1",
            story=cls.story,
        )
        ElementResultsCache.objects.create(
            project=cls.project,
            result_set=cls.result_set,
            result_type="WallShears_V2",
            element=cls.wall,
            story=cls.story,
            results_matrix={"TH01": 120.5, "TH02": 135.0},
            avg_value=127.75,
            max_value=135.0,
            min_value=120.5,
            load_case_count=2,
            story_sort_order=1,
        )

    def test_headers_and_rows(self):
        service = ResultDataService(self.project)
        result = _get_element_export_table_rows(
            service, self.result_set, "WallShears_V2", "WallShears", include_summary=True
        )
        self.assertIsNotNone(result)
        headers, rows = result
        self.assertIn("Element", headers)
        self.assertIn("Story", headers)

    def test_none_when_empty(self):
        service = ResultDataService(self.project)
        result = _get_element_export_table_rows(
            service, self.empty_result_set, "WallShears_V2", "WallShears", include_summary=True
        )
        self.assertIsNone(result)


# ------------------------------------------------------------------
# generate_excel_export
# ------------------------------------------------------------------


class GenerateExcelExportTests(_ExportTaskTestMixin, TestCase):
    def test_returns_bytesio(self):
        output = generate_excel_export(
            self.project, self.result_set, ["Drifts"], ["X"], True
        )
        self.assertIsInstance(output, BytesIO)
        # Should be a valid xlsx (starts with PK zip header)
        output.seek(0)
        self.assertEqual(output.read(2), b"PK")

    def test_sheet_per_type_direction(self):
        from openpyxl import load_workbook

        output = generate_excel_export(
            self.project, self.result_set, ["Drifts"], ["X", "Y"], True
        )
        output.seek(0)
        wb = load_workbook(output)
        # At least Drifts_X should be present (Drifts_Y may be "No Data" if missing)
        self.assertIn("Drifts_X", wb.sheetnames)

    def test_progress_callback(self):
        callback = MagicMock()
        generate_excel_export(
            self.project, self.result_set, ["Drifts"], ["X"], True,
            progress_callback=callback,
        )
        self.assertTrue(callback.called)

    def test_element_types(self):
        elem = Element.objects.create(
            project=self.project, element_type="Wall", name="W-EXL", unique_name="W-EXL",
            story=self.story,
        )
        ElementResultsCache.objects.create(
            project=self.project,
            result_set=self.result_set,
            result_type="WallShears_V2",
            element=elem,
            story=self.story,
            results_matrix={"TH01": 100.0},
            avg_value=100.0,
            load_case_count=1,
            story_sort_order=1,
        )
        output = generate_excel_export(
            self.project, self.result_set, [], [], True,
            element_types=["WallShears"],
        )
        self.assertIsInstance(output, BytesIO)

    def test_joint_types(self):
        JointResultsCache.objects.create(
            project=self.project,
            result_set=self.result_set,
            result_type="SoilPressures_Min",
            shell_object="FDN-EXL",
            unique_name="FDN-EXL-A",
            results_matrix={"TH01": -200.0},
            avg_value=-200.0,
            load_case_count=1,
        )
        output = generate_excel_export(
            self.project, self.result_set, [], [], True,
            joint_types=["SoilPressures"],
        )
        self.assertIsInstance(output, BytesIO)


# ------------------------------------------------------------------
# generate_csv_export
# ------------------------------------------------------------------


class GenerateCsvExportTests(_ExportTaskTestMixin, TestCase):
    def test_returns_string(self):
        output = generate_csv_export(
            self.project, self.result_set, ["Drifts"], ["X"], True
        )
        self.assertIsInstance(output, str)

    def test_section_separators(self):
        output = generate_csv_export(
            self.project, self.result_set, ["Drifts"], ["X"], True
        )
        self.assertIn("--- Drifts X ---", output)

    def test_progress_callback(self):
        callback = MagicMock()
        generate_csv_export(
            self.project, self.result_set, ["Drifts"], ["X"], True,
            progress_callback=callback,
        )
        self.assertTrue(callback.called)


# ------------------------------------------------------------------
# process_export_job task
# ------------------------------------------------------------------


class ProcessExportJobTaskTests(_ExportTaskTestMixin, TestCase):
    """process_export_job is ``bind=True`` so Celery passes ``self`` automatically.

    With ``CELERY_TASK_ALWAYS_EAGER=True`` (test settings), calling
    ``process_export_job.delay(job_id)`` or ``process_export_job.apply((job_id,))``
    executes synchronously and Celery injects the task instance as ``self``.
    """

    @patch("apps.exporter.tasks.send_export_progress")
    @patch("apps.exporter.tasks.send_export_complete")
    @patch("apps.exporter.tasks.send_export_error")
    def test_status_completed(self, mock_error, mock_complete, mock_progress):
        job = ExportJob.objects.create(
            project=self.project,
            user=self.user,
            export_format="excel",
            export_config={
                "result_set_id": self.result_set.id,
                "result_types": ["Drifts"],
                "directions": ["X"],
                "include_summary": True,
            },
        )
        process_export_job.apply(args=(job.id,))

        job.refresh_from_db()
        self.assertEqual(job.status, "completed")
        mock_complete.assert_called_once()

    @patch("apps.exporter.tasks.send_export_progress")
    @patch("apps.exporter.tasks.send_export_complete")
    @patch("apps.exporter.tasks.send_export_error")
    def test_failed_on_missing_result_set(self, mock_error, mock_complete, mock_progress):
        job = ExportJob.objects.create(
            project=self.project,
            user=self.user,
            export_format="excel",
            export_config={
                "result_set_id": 99999,
                "result_types": ["Drifts"],
                "directions": ["X"],
            },
        )
        process_export_job.apply(args=(job.id,))

        job.refresh_from_db()
        self.assertEqual(job.status, "failed")

    @patch("apps.exporter.tasks.send_export_progress")
    @patch("apps.exporter.tasks.send_export_complete")
    @patch("apps.exporter.tasks.send_export_error")
    def test_saves_output_file(self, mock_error, mock_complete, mock_progress):
        job = ExportJob.objects.create(
            project=self.project,
            user=self.user,
            export_format="csv",
            export_config={
                "result_set_id": self.result_set.id,
                "result_types": ["Drifts"],
                "directions": ["X"],
                "include_summary": True,
            },
        )
        process_export_job.apply(args=(job.id,))

        job.refresh_from_db()
        self.assertEqual(job.status, "completed")
        self.assertTrue(job.file_name.endswith(".csv"))
        self.assertGreater(job.file_size, 0)
